"""Shared Excel helper utilities."""

from __future__ import annotations

from pathlib import Path
from typing import Sequence, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _require_xlwings():
    try:
        import xlwings as xw  # type: ignore
    except ImportError as exc:
        raise ImportError(
            "OpenExcel 依赖 xlwings。请安装可选依赖: pip install wei-data-shu[excel-client]"
        ) from exc
    return xw


def create_workbook(file_path: Union[str, Path], default_sheet: str = "sheet1") -> None:
    if not default_sheet or not isinstance(default_sheet, str):
        raise ValueError("工作表名称必须是有效的字符串")

    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet is not None:
            sheet.title = default_sheet
        else:
            wb.create_sheet(title=default_sheet)
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        wb.close()
    except Exception as exc:
        raise IOError(f"创建工作簿失败: {exc}") from exc


def _auto_range(
    start_row: int,
    start_col: int,
    data: Sequence[Sequence],
    use_explicit: int,
    end_row: int,
    end_col: int,
) -> Tuple[int, int]:
    if use_explicit == 0 and data and len(data) > 0:
        calculated_end_row = len(data) + start_row - 1
        calculated_end_col = len(data[0]) + start_col - 1 if data[0] else start_col
        return calculated_end_row, calculated_end_col
    return end_row, end_col


def _apply_styles(
    worksheet: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    header_style: bool = True,
) -> None:
    font = Font(name="Microsoft YaHei", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    align = Alignment(vertical="center", horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = header_font if (header_style and row == start_row) else font
            cell.alignment = align
            cell.border = thin_border
            if header_style and row == start_row:
                cell.fill = header_fill

    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(start_row, end_row + 1):
            val = worksheet.cell(row=row, column=col).value
            length = len(str(val)) if val is not None else 0
            if length > max_len:
                max_len = length
        worksheet.column_dimensions[letter].width = max(8, min(int(max_len * 1.2) + 2, 50))


__all__ = ["create_workbook"]
