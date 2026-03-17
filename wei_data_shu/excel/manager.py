"""Core workbook manager."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ._helpers import _apply_styles, _auto_range, create_workbook


class ExcelManager:
    def __init__(self, file_path: Union[str, Path, None], default_sheet: str = "sheet1"):
        if file_path is None:
            raise ValueError("file_path 不能为 None")
        self.file_path = Path(file_path)
        self._workbook: Optional[Workbook] = None
        if not self.file_path.parent.exists():
            raise FileNotFoundError(f"目录不存在: {self.file_path.parent}")
        try:
            if not self.file_path.exists():
                create_workbook(str(self.file_path), default_sheet)
            self._workbook = load_workbook(str(self.file_path))
        except Exception as exc:
            raise IOError(f"加载工作簿失败: {exc}") from exc

    def __enter__(self) -> "ExcelManager":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        if exc_type is None:
            self.save()
        self.close()

    @property
    def workbook(self) -> Workbook:
        if self._workbook is None:
            raise RuntimeError("工作簿已关闭")
        return self._workbook

    @property
    def sheet_names(self) -> List[str]:
        return self.workbook.sheetnames

    def _ensure_sheet(self, sheet_name: str) -> Worksheet:
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
        return self.workbook[sheet_name]

    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> Worksheet:
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 已存在")
        return self.workbook.create_sheet(title=sheet_name, index=index)

    def delete_sheet(self, sheet_name: str) -> None:
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        if len(self.workbook.sheetnames) == 1:
            raise ValueError("不能删除唯一的工作表")
        self.workbook.remove(self.workbook[sheet_name])

    def write_sheet(
        self,
        sheet_name: str,
        data: Sequence[Sequence],
        start_row: int = 1,
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
        apply_styles: bool = True,
        header_row: bool = True,
    ) -> None:
        if not data:
            raise ValueError("数据不能为空")
        if end_row is None:
            end_row = start_row + len(data) - 1
        if end_col is None:
            max_cols = max(len(row) for row in data) if data else 0
            end_col = start_col + max_cols - 1

        worksheet = self._ensure_sheet(sheet_name)
        for i, row_data in enumerate(data):
            row_idx = start_row + i
            for j, value in enumerate(row_data):
                worksheet.cell(row=row_idx, column=start_col + j, value=value)

        if apply_styles:
            _apply_styles(worksheet, start_row, start_col, end_row, end_col, header_row)

    def read_sheet(
        self,
        sheet_name: str,
        start_row: int = 1,
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
    ) -> List[List[Any]]:
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        worksheet = self.workbook[sheet_name]
        if end_row is None:
            end_row = worksheet.max_row
        if end_col is None:
            end_col = worksheet.max_column

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(worksheet.cell(row=row, column=col).value)
            data.append(row_data)
        return data

    def fast_write(
        self,
        sheet_name: str,
        data: Sequence[Sequence],
        start_row: int = 1,
        start_col: int = 1,
        use_explicit_range: bool = False,
        end_row: int = 0,
        end_col: int = 0,
        header_row: bool = True,
    ) -> None:
        actual_end_row, actual_end_col = _auto_range(
            start_row, start_col, data, 1 if use_explicit_range else 0, end_row, end_col
        )
        self.write_sheet(
            sheet_name,
            data,
            start_row,
            start_col,
            actual_end_row,
            actual_end_col,
            apply_styles=True,
            header_row=header_row,
        )

    def write_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        start_row: int = 1,
        start_col: int = 1,
        include_header: bool = True,
        index: bool = False,
    ) -> None:
        data = []
        if include_header:
            headers = list(df.columns)
            if index:
                headers = [df.index.name or ""] + headers
            data.append(headers)
        for idx, row in df.iterrows():
            row_data = list(row)
            if index:
                row_data = [idx] + row_data
            data.append(row_data)
        self.write_sheet(sheet_name, data, start_row, start_col, apply_styles=True, header_row=include_header)

    def read_dataframe(self, sheet_name: str, start_row: int = 1, header_row: int = 1) -> pd.DataFrame:
        data = self.read_sheet(sheet_name, start_row=start_row)
        if not data:
            return pd.DataFrame()
        headers = [str(h) for h in data[0]]
        rows = data[1:] if len(data) > 1 else []
        return pd.DataFrame(rows, columns=headers)  # type: ignore

    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        worksheet = self.workbook[sheet_name]
        return {
            "name": sheet_name,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "dimensions": worksheet.dimensions,
            "index": self.workbook.sheetnames.index(sheet_name),
        }

    def copy_sheet(self, source_name: str, target_name: str) -> Worksheet:
        if source_name not in self.workbook.sheetnames:
            raise ValueError(f"源工作表 '{source_name}' 不存在")
        if target_name in self.workbook.sheetnames:
            raise ValueError(f"目标工作表 '{target_name}' 已存在")
        source = self.workbook[source_name]
        copied = self.workbook.copy_worksheet(source)
        copied.title = target_name
        return copied

    def save(self, file_path: Optional[Union[str, Path]] = None) -> None:
        save_path = file_path or self.file_path
        try:
            Path(save_path).parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(save_path)
        except Exception as exc:
            raise IOError(f"保存工作簿失败: {exc}") from exc

    def close(self) -> None:
        if self._workbook:
            try:
                self._workbook.close()
            except Exception:
                pass
            finally:
                self._workbook = None

    @classmethod
    def create(
        cls,
        file_path: Union[str, Path],
        default_sheet: str = "sheet1",
        overwrite: bool = False,
    ) -> "ExcelManager":
        path = Path(file_path)
        if path.exists() and not overwrite:
            raise FileExistsError(f"文件已存在: {path}")
        if path.exists():
            path.unlink()
        return cls(file_path, default_sheet)

    @classmethod
    def quick(cls, file_path: Union[str, Path], default_sheet: str = "sheet1") -> "ExcelManager":
        return cls(file_path, default_sheet)


eExcel = ExcelManager

__all__ = ["ExcelManager", "eExcel"]
